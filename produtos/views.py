from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import TipoProdutoForm, ItemPedidoForm, Loteform, AdicionarEstoqueForm, BaixaEstoqueForm
from usuarios.views import is_nutricionista, is_secretario_or_nutricionista
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Pedido, ItemPedido, TipoProduto, Lote, Usuario
from django.db import transaction
from django.db.models import Q, F, Sum, IntegerField, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from django.contrib import messages
from datetime import date, datetime, timedelta
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Apenas nutricionista ou secretário pode ver a lista de produtos do estoque
@login_required
@user_passes_test(is_secretario_or_nutricionista) 
def listar_tipos_produto(request):
    tipos_produto = TipoProduto.objects.all()
    
    busca = request.GET.get('buscar')
    if busca:
        tipos_produto = tipos_produto.filter(nome__icontains=busca)
        
    tipo = request.GET.get('tipo')
    if tipo and tipo != 'TODOS':
        tipos_produto = tipos_produto.filter(tipo=tipo)
        
    ordem = request.GET.get('ordem')
    if ordem == 'nome':
        tipos_produto = tipos_produto.order_by('nome')
        
    ativo = request.GET.get('ativo')
    if ativo == 'False':
        tipos_produto = tipos_produto.filter(ativo=False)
    else:
        tipos_produto = tipos_produto.filter(ativo=True)
    
    return render(request, 'produtos/listar_tipos.html', {
        'tipos_produto': tipos_produto,
        'tipos': TipoProduto.PRODUTO_TIPO,
        'busca': busca or '',
        'tipo_selecionado': tipo or 'TODOS',
        'ordem': ordem or '',
        'ativo': ativo or 'True',
    }) 



# Apenas nutricionista pode cadastrar produtos no estoque
@login_required
@user_passes_test(is_secretario_or_nutricionista)
def cadastrar_tipo_produto(request):
    if request.method == 'POST':
        form = TipoProdutoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_tipos_produto')
    else:
        form = TipoProdutoForm()
    
    return render(request, 'produtos/cadastrar_tipo.html', {'form': form})


# Apenas nutricionista ou secretário pode listar lotes de produtos no estoque
@login_required
@user_passes_test(is_secretario_or_nutricionista)
def listar_lotes(request, tipo_produto_id):
    tipo_produto = get_object_or_404(TipoProduto, pk=tipo_produto_id)
    lotes = Lote.objects.filter(tipo_produto=tipo_produto).order_by(F('data_validade').asc(nulls_last=True))
    
    hoje = date.today()
    alerta_dias = 30
    
    for lote in lotes:
        if lote.data_validade is None:
            lote.alerta_classe = 'lote-sem-validade' 
            continue 

        if lote.data_validade < hoje:
            lote.alerta_classe = 'lote-vencido'
        elif lote.data_validade <= hoje + timedelta(days=alerta_dias):
            lote.alerta_classe = 'lote-proximo-vencimento'
        else:
            lote.alerta_classe = ''
            
    context = {'tipo_produto': tipo_produto, 'lotes': lotes}
    return render(request, 'produtos/listar_lotes.html', context)
    


# Apenas nutricionista pode cadastrar lotes de produtos no estoque
@login_required
@user_passes_test(is_nutricionista)
def cadastrar_lote(request, tipo_produto_id):
    tipo_produto = get_object_or_404(TipoProduto, id=tipo_produto_id)

    if request.method == 'POST':
        form = Loteform(request.POST, tipo_produto=tipo_produto)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.tipo_produto = tipo_produto 
            lote.save()
            return redirect('listar_lotes', tipo_produto_id=tipo_produto.id)
    else:
        form = Loteform(tipo_produto=tipo_produto)
    
    context = {
        'form': form,
        'tipo_produto': tipo_produto,
    }
    return render(request, 'produtos/cadastrar_lote.html', context)



# Apenas nutricionista pode editar tipos de produtos no estoque
@login_required
@user_passes_test(is_nutricionista)
def editar_tipo_produto(request, tipo_produto_id):
    tipo_produto = get_object_or_404(TipoProduto, pk=tipo_produto_id)
    
    if request.method == 'POST':
        form = TipoProdutoForm(request.POST, instance=tipo_produto)
        if form.is_valid():
            form.save()
            return redirect('listar_tipos_produto')
    else:
        form = TipoProdutoForm(instance=tipo_produto)
        
    return render(request, 'produtos/cadastrar_tipo.html', {
        'form': form, 
        'editar': True
    })
    
    

# Apenas nutricionista pode excluir tipos de produtos no estoque
@login_required
@user_passes_test(is_nutricionista)
def excluir_tipo_produto(request, tipo_produto_id):
    tipo_produto = get_object_or_404(TipoProduto, pk=tipo_produto_id)
    
    if request.method == 'POST':
        tipo_produto.ativo = False
        tipo_produto.save()
        messages.success(request, f"Produto '{tipo_produto.nome}' foi arquivado com sucesso.")
        return redirect('listar_tipos_produto')

    return render(request, 'produtos/excluir_tipo.html', {'tipo_produto': tipo_produto}) 



# Apenas nutricionista pode reativar um produto desativado
@login_required
@user_passes_test(is_nutricionista)
def reativar_tipo_produto(request, tipo_produto_id):
    tipo_produto = get_object_or_404(TipoProduto, pk=tipo_produto_id)
    
    if request.method == 'POST':
        tipo_produto.ativo = True
        tipo_produto.save()
        messages.success(request, f"Produto '{tipo_produto.nome}' foi reativado com sucesso.")
        return redirect('listar_tipos_produto')
    
    return render(request, 'produtos/reativar_tipo.html', {'tipo_produto': tipo_produto})


# Apenas nutricionista pode adicionar produto no estoque
@login_required
@user_passes_test(is_nutricionista)
def adicionar_estoque(request, lote_id):
    lote = get_object_or_404(Lote, pk=lote_id)
    
    if request.method == 'POST':
        form = AdicionarEstoqueForm(request.POST) 
        
        if form.is_valid():
            novos_pacotes = form.cleaned_data['quantidade_pacotes'] 
            lote.quantidade_pacotes += novos_pacotes
            lote.save()
            
            return redirect('listar_lotes', tipo_produto_id=lote.tipo_produto.id)
    else:
        form = AdicionarEstoqueForm()
    
    return render(request, 'produtos/adicionar_estoque.html', {'form': form, 'lote': lote}) 


# Apenas nutricionista pode baixar produtos do estoque
@login_required
@user_passes_test(is_nutricionista)
def baixar_estoque(request, lote_id):
    lote = get_object_or_404(Lote, pk=lote_id)
    
    if request.method == 'POST':
        form = BaixaEstoqueForm(request.POST) 
        
        if form.is_valid():
            quantidade_remover = form.cleaned_data['quantidade_pacotes']
            
            if quantidade_remover > lote.quantidade_pacotes:
                form.add_error('quantidade', "Não há unidades suficientes no estoque.")
            else:
                lote.quantidade_pacotes -= quantidade_remover
                lote.save()
                return redirect('listar_lotes', tipo_produto_id=lote.tipo_produto.id)   
    else:      
        form = BaixaEstoqueForm() 
    
    return render(request, 'produtos/baixar_estoque.html', {
        'lote': lote,
        'form': form, 
    })



# Apenas nutricionista pode editar lotes do estoque
@login_required
@user_passes_test(is_nutricionista)
def editar_lote(request, lote_id):
    lote = get_object_or_404(Lote, pk=lote_id)
    
    if request.method == 'POST':
        form = Loteform(request.POST, instance=lote)
        if form.is_valid():
            form.save()
            return redirect('listar_lotes', tipo_produto_id=lote.tipo_produto.id)
    else:
        form = Loteform(instance=lote)
        
    return render(request, 'produtos/cadastrar_lote.html', {
        'form': form, 
        'editar': True,
        'lote': lote,
        'tipo_produto': lote.tipo_produto,
    })



# Apenas nutricionista pode excluir lotes do estoque
@login_required
@user_passes_test(is_nutricionista)
def excluir_lote(request, lote_id):
    lote = get_object_or_404(Lote, pk=lote_id)
    
    if request.method == 'POST':
        lote.delete()
        return redirect('listar_lotes', tipo_produto_id=lote.tipo_produto.id)
    
    return render(request, 'produtos/excluir_lote.html', {'lote': lote})



def exportar_excel(dados, mes, ano, usuario, total_pedidos, total_pedidos_carne, lista_detalhada_carne):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Consumo"

    ws['A1'] = f"Balanço de Consumo por Unidade: {usuario}"
    ws['A1'].font = Font(bold=True, size=14)
    
    mes_str = f"{mes}/{ano}" if mes != '0' else f"Anual {ano}"
    ws['A2'] = f"Referência: {mes_str}"
    ws['A3'] = f"Total de Pedidos: {total_pedidos}"
    ws['A4'] = f"Pedidos com Carne: {total_pedidos_carne}"

    headers = ['Produto', 'Unidade', 'Qtd Total', 'Preço Unit. (R$)', 'Subtotal (R$)']
    header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    row_num = 7
    investimento_total = 0

    for item in dados:
        v_unit = float(item.get('preco_unitario_medio', 0) or 0)
        v_total = float(item.get('valor_total_produto', 0) or 0)
        investimento_total += v_total

        ws.cell(row=row_num, column=1).value = item['produto__nome']
        ws.cell(row=row_num, column=2).value = item['produto__unidade_medida']
        ws.cell(row=row_num, column=3).value = item['total_solicitado']
        ws.cell(row=row_num, column=4).value = float(v_unit)
        ws.cell(row=row_num, column=5).value = float(v_total)
        
        ws.cell(row=row_num, column=1).value = item['produto__nome']
        ws.cell(row=row_num, column=2).value = item['produto__unidade_medida']
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row_num, column=3).value = item['total_solicitado']
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        
        # Colunas de Dinheiro
        c_unit = ws.cell(row=row_num, column=4)
        c_unit.value = v_unit
        c_unit.number_format = '#,##0.00' 

        c_total = ws.cell(row=row_num, column=5)
        c_total.value = v_total
        c_total.number_format = '#,##0.00'
        
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=4).value = "INVESTIMENTO TOTAL:"
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
    
    total_cell = ws.cell(row=row_num, column=5)
    total_cell.value = investimento_total
    total_cell.font = Font(bold=True)
    total_cell.number_format = 'R$ #,##0.00' 

    if lista_detalhada_carne:
        row_num += 3
        ws.cell(row=row_num, column=1).value = "Detalhamento: Observações de Proteína (Carne)"
        ws.cell(row=row_num, column=1).font = Font(bold=True, underline="single", color="990000")
        
        row_num += 1
        for obs in lista_detalhada_carne:
            if obs:
                ws.cell(row=row_num, column=1).value = f"• {obs}"
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                row_num += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_consumo_{usuario}_{mes if mes else 'anual'}_{ano}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response



def exportar_pdf(dados, mes, ano, usuario, total_pedidos, total_pedidos_carne, lista_detalhada_carne, grafico=None):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_consumo_{usuario}_{mes}_{ano}.pdf'
    
    p = canvas.Canvas(response, pagesize=A4)
    p.setTitle(f"Relátorio_de_consumo_{usuario}_{mes}/{ano}")
    width, height = A4
    
    mes_int = int(mes) if mes else 0
    titulo_periodo = f"{mes}/{ano}" if mes_int > 0 else f"Anual {ano}"
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 800, "Balanço de Consumo por Unidade") 
    p.setFont("Helvetica", 11)
    p.drawString(50, 777, f"Unidade Solicitante: {usuario}")
    p.drawString(50, 760, f"Referência: {titulo_periodo}")
    
    p.setStrokeColor(colors.black)
    p.rect(50, 690, 500, 60, stroke=1, fill=0) 
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(60, 730, f"Total de Pedidos Realizados: {total_pedidos}")
    
    p.setFillColor(colors.dodgerblue)
    p.drawString(60, 707, f"Pedidos contendo Proteína (Carne): {total_pedidos_carne}")
    p.setFillColor(colors.black)
    
    y = 660
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, y, "Produto")
    p.drawString(280, y, "Unid")
    p.drawString(330, y, "Qtd")
    p.drawString(390, y, "V. Unit (R$)") 
    p.drawString(480, y, "Total (R$)")   
    
    p.setStrokeColor(colors.lightgrey)
    p.line(50, y-5, 550, y-5)
    
    y -= 25
    p.setFont("Helvetica", 10)
    
    valor_geral_relatorio = 0 

    for item in dados:
        p.drawString(50, y, item['produto__nome'][:40])
        
        p.drawString(280, y, item['produto__unidade_medida'])
        
        p.drawString(330, y, f"{item['total_solicitado']:,}")
        
        v_unit = item.get('preco_unitario_medio', 0) or 0
        p.drawString(390, y, f"{v_unit:,.2f}")
        
        v_total = item.get('valor_total_produto', 0) or 0
        p.drawString(480, y, f"{v_total:,.2f}")
        
        valor_geral_relatorio += float(v_total)
        
        y -= 20
        
        if y < 80:
            p.showPage()
            y = 800
            p.setFont("Helvetica-Bold", 11)
            p.drawString(50, y, "Produto (cont.)")
            p.drawString(280, y, "Unid")
            p.drawString(330, y, "Qtd")
            p.drawString(390, y, "V. Unit")
            p.drawString(480, y, "Total")
            y -= 25
            p.setFont("Helvetica", 10)

    p.setStrokeColor(colors.black)
    p.line(350, y, 550, y)
    y -= 15
    p.setFont("Helvetica-Bold", 10)
    p.drawString(365, y, "INVESTIMENTO TOTAL:")
    p.drawString(490, y, f"R$ {valor_geral_relatorio:,.2f}")

    if total_pedidos_carne > 0:
        if y < 150: 
            p.showPage()
            y = 800
        
        y -= 40
        p.setFillColor(colors.darkred)
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Detalhamento: Observações de Proteína (Carne)")
        p.setFillColor(colors.black)
        p.line(50, y-5, 330, y-5)
        
        y -= 20
        p.setFont("Helvetica-Oblique", 9)
        for texto_carne in lista_detalhada_carne:
            if texto_carne: 
                texto_formatado = f"• {texto_carne[:110]}" 
                p.drawString(60, y, texto_formatado)
                y -= 15
                if y < 60:
                    p.showPage()
                    y = 800
    
    if grafico:
        img = ImageReader(grafico)
        altura_grafico = 160
        
        if y < 180:
            p.showPage()
            y = 800
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Análise de Investimento (Cont.)")
            y -= 30
        else:
            y -= 40

        p.drawImage(img, 50, y - altura_grafico, width=480, height=altura_grafico, preserveAspectRatio=True)
        y -= (altura_grafico + 20)
    
    p.save()
    return response



# Apenas secretário pode filtrar relatórios
@login_required
@user_passes_test(lambda u: u.is_secretario())
def filtros_relatorio_mensal(request):
    ano_atual = datetime.now().year
    anos_disponiveis = range(2026, ano_atual + 1)
    
    meses_lista = [
        ('0', 'Anual'), ('1', 'Janeiro'), ('2', 'Fevereiro'), ('3', 'Março'),
        ('4', 'Abril'), ('5', 'Maio'), ('6', 'Junho'), ('7', 'Julho'),
        ('8', 'Agosto'), ('9', 'Setembro'), ('10', 'Outubro'), ('11', 'Novembro'), ('12', 'Dezembro')
    ]
    
    diretores = Usuario.objects.filter(cargo='DIRETOR')

    context = {
        'anos': anos_disponiveis,
        'ano_atual': ano_atual,
        'meses': meses_lista,
        'diretores': diretores,
    }
    return render(request, 'produtos/filtros_relatorio.html', context)
    
    
    
def exportar_excel_origemProdutos(produtos_agri, produtos_comum, total_geral, total_agri, total_comum, porcentagem, mes, ano):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balanço de Compras"

    estilo_titulo = Font(bold=True, size=14)
    estilo_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento_azul = PatternFill(start_color="FF1E90FF", end_color="FF1E90FF", fill_type="solid")
    preenchimento_verde = PatternFill(start_color="FF228B22", end_color="FF228B22", fill_type="solid")
    alinhamento_centro = Alignment(horizontal="center")
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = "Balanço de Compras: Origem Produto"
    ws['A1'].font = estilo_titulo
    ws['A2'] = f"Referência: {mes}/{ano}"
    
    ws['A4'] = "CATEGORIA"
    ws['B4'] = "PORCENTAGEM (%)"
    ws['C4'] = "TOTAL INVESTIDO (R$)"
    for cell in ws[4]: 
        cell.font = Font(bold=True)
        cell.border = borda_fina

    ws.append(["Compra Comum", f"{round((total_comum/total_geral*100), 2) if total_geral > 0 else 0}%", total_comum])
    ws.append(["Agricultura Familiar", f"{porcentagem}%", total_agri])
    ws.append(["GASTO TOTAL", "100%", total_geral])
    ws[ws.max_row][0].font = Font(bold=True)

    ws.append([]) 
    ws.append(["DETALHAMENTO: COMPRA COMUM"])
    ws[ws.max_row][0].font = Font(bold=True, color="1E90FF")
    
    colunas = ["Produto", "Qtd", "Unid", "Subtotal (R$)"]
    ws.append(colunas)
    for cell in ws[ws.max_row]:
        cell.fill = preenchimento_azul
        cell.font = estilo_cabecalho
        cell.alignment = alinhamento_centro

    for item in produtos_comum:
        ws.append([item['tipo_produto__nome'], item['qtd'], item['tipo_produto__unidade_medida'], item['subtotal'] or 0])

    ws.append([])
    ws.append(["DETALHAMENTO: AGRICULTURA FAMILIAR"])
    ws[ws.max_row][0].font = Font(bold=True, color="228B22")
    
    ws.append(colunas)
    for cell in ws[ws.max_row]:
        cell.fill = preenchimento_verde
        cell.font = estilo_cabecalho
        cell.alignment = alinhamento_centro

    for item in produtos_agri:
        ws.append([item['tipo_produto__nome'], item['qtd'], item['tipo_produto__unidade_medida'], item['subtotal'] or 0])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=relatorio_compras_{mes}_{ano}.xlsx'
    wb.save(response)
    return response 
    
    
    
def exportar_excel_tipos(dados, total_valor, total_produtos, total_lotes, mes, ano):
    wb = openpyxl.Workbook()
    
    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Categoria"
    ws_resumo.append(["Balanço por Categoria", f"{mes}/{ano}"])
    ws_resumo.append(["Total de Produtos", total_produtos])
    ws_resumo.append(["Total de Itens", total_lotes])
    ws_resumo.append(["Investimento Total", total_valor])
    
    categorias_config = [
        ('ALIMENTO', 'Alimentos', "228B22"),   
        ('LIMPEZA', 'Limpeza', "1E90FF"),      
        ('ESCOLAR', 'Escolar', "FF8C00"),      
        ('OUTROS', 'Outros', "808080"),       
    ]

    for chave, label, cor_hex in categorias_config:
        info = dados[chave]
        if info['lista']:
            ws = wb.create_sheet(title=label)
            ws.sheet_properties.tabColor = cor_hex
            
            ws.append([f"DETALHAMENTO: {label.upper()}"])
            ws.append(["Produto", "Quantidade", "Unidade", "Subtotal (R$)"])
            
            for item in info['lista']:
                ws.append([
                    item['tipo_produto__nome'], 
                    item['qtd'], 
                    item['tipo_produto__unidade_medida'], 
                    item['subtotal']
                ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=relatorio_categorias_{mes}_{ano}.xlsx'
    wb.save(response)
    return response  
    


def exportar_pdf_origemProdutos(produtos_agri, produtos_comum, total_geral, total_agri, total_comum, porcentagem, mes, ano):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_produtos_{mes}_{ano}.pdf'
    
    p = canvas.Canvas(response, pagesize=A4)
    p.setTitle(f"relatório_origem_produtos_{mes}/{ano}")
    titulo_periodo = f"{mes}/{ano}" if mes > 0 else f"Anual {ano}"
    
    porcentagem_comum = round((total_comum / total_geral * 100), 2) if total_geral > 0 else 0
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Balanço de Compras: Origem Produto")
    p.setFont("Helvetica", 12)
    p.drawString(100, 775, f"Referência: {titulo_periodo}")
    
    p.setStrokeColor(colors.dodgerblue)
    p.rect(100, 705, 400, 60, stroke=1, fill=0) 
    
    p.setFont("Helvetica-Bold", 13)
    p.setFillColor(colors.dodgerblue)
    p.drawString(120, 740, f"Índice de Compra Comum: {porcentagem_comum}%")
    
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.black)
    p.drawString(120, 720, f"Gasto no Mercado/Atacado: R$ {total_comum:,.2f}")
    
    cor_meta = colors.green if porcentagem >= 15 else colors.red
    p.setStrokeColor(cor_meta)
    p.rect(100, 620, 400, 75, stroke=1, fill=0) 
    
    p.setFont("Helvetica-Bold", 13)
    p.setFillColor(cor_meta)
    p.drawString(120, 675, f"Índice de Agricultura Familiar: {porcentagem}%")
    
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.black)
    p.drawString(120, 655, f"Gasto com produto de Agricultura: R$ {total_agri:,.2f}")
    
    if porcentagem < 15:
        p.setFont("Helvetica-Oblique", 9)
        p.setFillColor(colors.red)
        p.drawString(120, 635, "* Atenção: Abaixo do limite legal de 15%")
    
    p.setFont("Helvetica-Bold", 11)
    p.drawCentredString(300, 595, f"GASTO TOTAL EM ALIMENTOS: R$ {total_geral:,.2f}")
    
    y = 565 
    
    def desenhar_cabecalho_tabela(y_pos, titulo, cor):
        p.setFillColor(cor)
        p.setFont("Helvetica-Bold", 11)
        p.drawString(100, y_pos, titulo)
        y_pos -= 20
        p.setFillColor(colors.black)
        p.drawString(100, y_pos, "Produto")
        p.drawString(300, y_pos, "Qtd")
        p.drawString(380, y_pos, "Unid")
        p.drawString(450, y_pos, "Subtotal (R$)")
        return y_pos - 15

    y = desenhar_cabecalho_tabela(y, "Detalhe: Compra Comum", colors.dodgerblue)
    p.setFont("Helvetica", 10)
    for item in produtos_comum:
        p.drawString(100, y, item['tipo_produto__nome'][:30])
        p.drawString(300, y, str(item['qtd']))
        p.drawString(380, y, item['tipo_produto__unidade_medida'])
        subtotal = item['subtotal'] or 0
        p.drawString(450, y, f"{subtotal:,.2f}")
        y -= 15
        if y < 60: p.showPage(); y = 800
        
    y -= 30 # Garante que as tabelas não enconstem uma na outra
    
    y = desenhar_cabecalho_tabela(y, "Detalhe: Agricultura Familiar", colors.green)
    p.setFont("Helvetica", 10)
    for item in produtos_agri:
        p.drawString(100, y, item['tipo_produto__nome'][:30])
        p.drawString(300, y, str(item['qtd']))
        p.drawString(380, y, item['tipo_produto__unidade_medida'])
        subtotal = item['subtotal'] or 0
        p.drawString(450, y, f"{subtotal:,.2f}")
        y -= 15
        if y < 60: p.showPage(); y = 800

    p.save()
    return response



def exportar_pdf_tipos(dados, total_valor, total_lotes, total_produtos, mes, ano):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_tipos_{mes}_{ano}.pdf'
    p = canvas.Canvas(response, pagesize=A4)
    p.setTitle(f"Relatório_tipos_produtos_{mes}/{ano}")
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 810, "Balanço de Compras por Categoria")
    p.setFont("Helvetica", 12)
    p.drawString(100, 787, f"Referência: {mes}/{ano} | Total Geral: R$ {total_valor:,.2f}")

    p.setStrokeColor(colors.black)
    p.rect(100, 710, 400, 65, stroke=1, fill=0)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(110, 760, f"Total de Produtos Cadastrados: {total_produtos}")
    p.drawString(110, 740, f"Total de Lotes Cadastrados (Todos os tipos): {total_lotes}")
    p.drawString(110, 720, f"Investimento Total no Período: R$ {total_valor:,.2f}")

    y = 680
    categorias_config = [
        ('ALIMENTO', 'Alimentos', colors.green),
        ('LIMPEZA', 'Materiais de Limpeza', colors.dodgerblue),
        ('ESCOLAR', 'Materiais Escolares', colors.orange),
        ('OUTROS', 'Outros/Diversos', colors.grey),
    ]

    for chave, label, cor in categorias_config:
        info = dados[chave]
        if info['lista']:
            p.setFillColor(cor)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, y, f"Tipo: {label} (Total: R$ {info['valor']:,.2f})")
            y -= 20
            
            p.setFillColor(colors.black)
            p.setFont("Helvetica-Bold", 10)
            p.drawString(100, y, "Produto")
            p.drawString(300, y, "Qtd")
            p.drawString(450, y, "Subtotal")
            y -= 15
            
            p.setFont("Helvetica", 9)
            for item in info['lista']:
                p.drawString(100, y, item['tipo_produto__nome'][:35])
                p.drawString(300, y, f"{item['qtd']} {item['tipo_produto__unidade_medida']}")
                subtotal = item['subtotal'] or 0
                p.drawString(450, y, f"R$ {subtotal:,.2f}")
                y -= 15
                if y < 60:
                    p.showPage()
                    y = 800
            y -= 20 

    p.save()
    return response



# Apenas secretário pode gerar relatório de consumo
@user_passes_test(lambda u: u.is_secretario())
def gerar_relatorio_mensal(request):
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    diretor_id = request.GET.get('diretor')
    formato = request.GET.get('formato')

    filtros_pedidos = {
        'data_pedido__year': ano,
        'status__in': ['PENDENTE', 'APROVADO', 'ENTREGUE']
    }
    
    filtros_itens = {
        'pedido__data_pedido__year': ano,
        'pedido__status__in': ['PENDENTE', 'APROVADO', 'ENTREGUE']
    }
    
    if mes and mes != '0':
        filtros_pedidos['data_pedido__month'] = mes
        filtros_itens['pedido__data_pedido__month'] = mes

    nome_diretor = "Todas as Escolas"
    if diretor_id and diretor_id != 'todos':
        filtros_pedidos['solicitante_id'] = diretor_id
        filtros_itens['pedido__solicitante_id'] = diretor_id
        
        diretor_obj = Usuario.objects.get(id=diretor_id)
        
        # Tenta pegar o nome da escola, se não tiver, usa o username mesmo
        if hasattr(diretor_obj, 'escola_diretoria'):
            nome_diretor = diretor_obj.escola.nome_escola
        else:
            nome_diretor = diretor_obj.username

    pedidos_no_periodo = Pedido.objects.filter(**filtros_pedidos)
    total_pedidos = pedidos_no_periodo.count()
    
    pedidos_carne = pedidos_no_periodo.exclude(observacao_carne="").exclude(observacao_carne__isnull=True).select_related('solicitante', 'solicitante__escola')
    total_pedidos_carne = pedidos_carne.count()
    is_todas_escolas = (diretor_id == 'todos' or not diretor_id)
    
    lista_detalhada_carne = []
    for p in pedidos_carne:
        if is_todas_escolas:
            # Verifica se existe o perfil escola para exibir o nome correto
            if hasattr(p.solicitante, 'escola'):
                nome_exibicao = p.solicitante.escola.nome_escola
            else:
                nome_exibicao = p.solicitante.username
            
            lista_detalhada_carne.append(f"{nome_exibicao}: {p.observacao_carne}")
        else:
            lista_detalhada_carne.append(p.observacao_carne)
    
    
    preco_real_subquery = Lote.objects.filter(
        tipo_produto_id=OuterRef('produto_id')
    ).order_by('-id').values('preco_unitario')[:1]

    dados_raw = ItemPedido.objects.filter(**filtros_itens).values(
        'produto__nome',
        'produto__unidade_medida',
        'produto_id'
    ).annotate(
        total_solicitado=Sum('quantidade', distinct=True),
        valor_total_produto=Sum(
            F('quantidade') * Subquery(preco_real_subquery),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('produto__nome')
    
    dados = []
    for item in dados_raw:
        qtd = item.get('total_solicitado') or 0
        total_rs = item.get('valor_total_produto') or 0
    
        dados.append({
            'produto__nome': item['produto__nome'],
            'produto__unidade_medida': item['produto__unidade_medida'],
            'total_solicitado': qtd,
            'preco_unitario_medio': float(total_rs / qtd) if qtd > 0 else 0,
            'valor_total_produto': float(total_rs)
        })

    grafico_buffer = None
    if diretor_id == 'todos' or not diretor_id:
        dados_grafico = ItemPedido.objects.filter(**filtros_itens).values(
            'pedido__solicitante__username',
            'pedido__solicitante__escola__nome_escola'
        ).annotate(
            total_escola=Sum(
                F('quantidade') * Subquery(preco_real_subquery),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        ).order_by('-total_escola')
    
        if dados_grafico.exists():
            grafico_buffer = gerar_grafico_consumo_escolas(dados_grafico)

    if formato == 'excel':
        return exportar_excel(dados, mes, ano, nome_diretor, total_pedidos, total_pedidos_carne, lista_detalhada_carne)
    else:
        return exportar_pdf(dados, mes, ano, nome_diretor, total_pedidos, total_pedidos_carne, lista_detalhada_carne, grafico_buffer)



def gerar_grafico_consumo_escolas(dados_grafico):
    
    escolas = []
    valores = []
    
    for item in dados_grafico:
        nome_oficial = item.get('pedido__solicitante__escola__nome_escola')
        username = item.get('pedido__solicitante__username')
        
        label = nome_oficial if nome_oficial else username
        escolas.append(label)
        valores.append(float(item['total_escola'] or 0))
    
    total_periodo = sum(valores) if sum(valores) > 0 else 1

    fig, ax = plt.subplots(figsize=(10, len(escolas) * 0.5 + 2))
    bars = ax.barh(escolas, valores, color='#1E90FF', height=0.6) 

    for bar in bars:
        largura = bar.get_width()
        porcentagem = (largura / total_periodo) * 100
        
        valor_fmt = f"{largura:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        porcentagem_fmt = f"{porcentagem:.1f}".replace(".", ",")
        
        label_texto = f"R$ {valor_fmt} ({porcentagem_fmt}%)"
        
        ax.text(largura + (total_periodo * 0.02), bar.get_y() + bar.get_height()/2,
                label_texto, va='center', fontweight='bold', fontsize=10, color='#333333') 

    ax.set_title('Divisão de Gastos por Unidade (%)', fontsize=12, pad=15)
    ax.xaxis.set_visible(False)
    
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    
    ax.invert_yaxis()
    ax.tick_params(axis='y', length=0)

    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight', dpi=120)
    plt.close(fig)
    buffer.seek(0)
    return buffer



# Apenas secretário pode gerar relatório de origem
@login_required
@user_passes_test(lambda u: u.is_secretario())
def gerar_relatorio_origem(request):
    mes = int(request.GET.get('mes', 0))
    ano = int(request.GET.get('ano', 2026))
    formato = request.GET.get('formato')

    filtros = {'data_cadastro__year': ano}
    if mes > 0:
        filtros['data_cadastro__month'] = mes

    lotes_base = Lote.objects.filter(tipo_produto__tipo='ALIMENTO', **filtros).select_related('tipo_produto')

    lotes_agri = lotes_base.filter(tipo_produto__origem_compra='AGRICULTURA')
    total_agri = lotes_agri.aggregate(total=Sum(F('quantidade_pacotes') * F('preco_unitario')))['total'] or 0
    
    lotes_comum = lotes_base.filter(tipo_produto__origem_compra='COMUM')
    total_comum = lotes_comum.aggregate(total=Sum(F('quantidade_pacotes') * F('preco_unitario')))['total'] or 0

    total_geral = total_agri + total_comum
    porcentagem = round((total_agri / total_geral * 100), 2) if total_geral > 0 else 0

    produtos_agri = lotes_agri.values('tipo_produto__nome', 'tipo_produto__unidade_medida').annotate(
        qtd=Sum('quantidade_pacotes'), subtotal=Sum(F('quantidade_pacotes') * F('preco_unitario'))
    )
    
    produtos_comum = lotes_comum.values('tipo_produto__nome', 'tipo_produto__unidade_medida').annotate(
        qtd=Sum('quantidade_pacotes'), subtotal=Sum(F('quantidade_pacotes') * F('preco_unitario'))
    )

    if formato == 'pdf':
        return exportar_pdf_origemProdutos(produtos_agri, produtos_comum, total_geral, total_agri, total_comum, porcentagem, mes, ano)
    else:
        return exportar_excel_origemProdutos(produtos_agri, produtos_comum, total_geral, total_agri, total_comum, porcentagem, mes, ano)



# Apenas secretário pode gera relatório de tipo
@login_required
@user_passes_test(lambda u: u.is_secretario())
def gerar_relatorio_tipos(request):
    mes = int(request.GET.get('mes', 0))
    ano = int(request.GET.get('ano', 2026))
    formato = request.GET.get('formato')
    
    filtros = {'data_cadastro__year': ano}
    if mes > 0:
        filtros['data_cadastro__month'] = mes
        
    lotes_base = Lote.objects.filter(**filtros).select_related('tipo_produto')
    total_produtosDistintos = lotes_base.values('tipo_produto').distinct().count()
    total_lotes = lotes_base.count()
    
    def dados_por_tipo(slug_tipo):
        queryset = lotes_base.filter(tipo_produto__tipo=slug_tipo)
        total_valor = queryset.aggregate(total=Sum(F('quantidade_pacotes') * F('preco_unitario')))['total'] or 0
        qtd_itens = queryset.aggregate(total=Sum('quantidade_pacotes'))['total'] or 0
        detalhes = list(queryset.values('tipo_produto__nome', 'tipo_produto__unidade_medida').annotate(
            qtd=Sum('quantidade_pacotes'),
            subtotal=Sum(F('quantidade_pacotes') * F('preco_unitario'))
        ))
        return {'valor': total_valor, 'qtd': qtd_itens, 'detalhes': detalhes, 'lista': detalhes}
    
    dados = {
        'ALIMENTO': dados_por_tipo('ALIMENTO'),
        'LIMPEZA': dados_por_tipo('LIMPEZA'),
        'ESCOLAR': dados_por_tipo('ESCOLAR'),
        'OUTROS': dados_por_tipo('OUTROS')
    }
    
    total_geral_valor = sum(d['valor'] for d in dados.values())
    total_geral_itens = sum(d['qtd'] for d in dados.values())
    total_lotes = lotes_base.count()
    
    if formato == 'pdf':
        return exportar_pdf_tipos(dados, total_geral_valor, total_geral_itens, total_lotes, mes, ano)
    else:
        return exportar_excel_tipos(dados, total_geral_valor, total_geral_itens, total_lotes, mes, ano)



# Apenas diretor pode fazer um pedido
@login_required
@user_passes_test(lambda u: u.is_diretor())
def criar_pedido(request):
    novo_pedido = Pedido.objects.create(solicitante=request.user)
    return redirect('detalhe_pedido', pedido_id=novo_pedido.id)



# Apenas um diretor pode ver os detalhes do seu pedido
@login_required
@user_passes_test(lambda u: u.is_diretor())
def detalhe_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, solicitante=request.user)
    
    produtos_validos = TipoProduto.objects.filter(ativo=True).annotate(
        total_estoque=Coalesce(
            Sum('lotes__quantidade_pacotes'),
            0,
            output_field=IntegerField()
        )
    ).filter(total_estoque__gt=0)
    
    if request.method == 'POST':
        form = ItemPedidoForm(request.POST)
        form.fields['produto'].queryset = produtos_validos
        
        if form.is_valid():
            produto_selecionado = form.cleaned_data['produto']
            quantidade_adicionar = form.cleaned_data['quantidade']
            
            estoque_atual = produtos_validos.get(pk=produto_selecionado.pk).total_estoque
            item_existente = ItemPedido.objects.filter(pedido=pedido, produto=produto_selecionado).first()
            quantidade_existente = item_existente.quantidade if item_existente else 0
            quantidade_final_desejada = quantidade_existente + quantidade_adicionar

            if quantidade_final_desejada > estoque_atual:
                messages.error(request, f"Estoque insuficiente para '{produto_selecionado.nome}'. "
                                        f"Disponível: {estoque_atual}, "
                                        f"Solicitado: {quantidade_final_desejada}.")
            else:
                if item_existente:
                    item_existente.quantidade += quantidade_adicionar
                    item_existente.save()
                    messages.success(request, f"Quantidade de '{produto_selecionado.nome}' atualizada.")
                else:
                    item = form.save(commit=False)
                    item.pedido = pedido
                    item.save()
                    messages.success(request, f"'{produto_selecionado.nome}' adicionado ao pedido.")
            
            return redirect('detalhe_pedido', pedido_id=pedido.id)
        
        else:
            form = ItemPedidoForm()
            form.fields['produto'].queryset = produtos_validos
    else:
        form = ItemPedidoForm()
        form.fields['produto'].queryset = produtos_validos
        
    itens_do_pedido = pedido.itens.all()

    context = {
        'pedido': pedido,
        'itens': itens_do_pedido,
        'form': form,
    }
    return render(request, 'pedidos/detalhe_pedido.html', context)



#Apenas diretor pode remover um item de um pedido respectivamente seu
@login_required
@user_passes_test(lambda u: u.is_diretor())
def remover_item_pedido(request, item_id):
    item = get_object_or_404(ItemPedido, id=item_id)

    pedido_id = item.pedido.id

    if item.pedido.solicitante == request.user:
        item.delete()

        return redirect('detalhe_pedido', pedido_id=pedido_id)
    else:
        return HttpResponseForbidden("Você não tem permissão para remover este item.")



# Apenas diretor pode ver seus próprios pedidos
@login_required
@user_passes_test(lambda u: u.is_diretor())
def meus_pedidos(request):
    
    pedidos_do_diretor = Pedido.objects.filter(
        solicitante=request.user,
        itens__isnull=False 
    ).distinct().order_by('-data_pedido')
    
    context = {
        'pedidos': pedidos_do_diretor
    }
    
    return render(request, 'pedidos/meus_pedidos.html', context)



# Apenas diretor pode finalizar um pedido e encaminha-lo a nutricionista
@login_required
@user_passes_test(lambda u: u.is_diretor())
def finalizar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        observacao = request.POST.get('obs_carne')
        
        if observacao:
            pedido.observacao_carne = observacao
            
        pedido.status = 'PENDENTE' 
        pedido.save()
        
        return redirect('meus_pedidos')

    return redirect('detalhe_pedido', pedido_id=pedido.id)



# Apenas diretor pode verificar o estoque disponível
@login_required
@user_passes_test(lambda u: u.is_diretor())
def listar_estoque_disponivel(request):

    produtos = TipoProduto.objects.filter(ativo=True).annotate(
        total_estoque=Coalesce(
            Sum('lotes__quantidade_pacotes'), 
            0,
            output_field=IntegerField()
        )
    ).order_by('nome')
    
    context = {
        'produtos': produtos
    }
    
    return render(request, 'produtos/estoque_disponivel.html', context)



# Apenas diretor pode excluir seu próprio pedido
@login_required
@user_passes_test(lambda u: u.is_diretor())
def excluir_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, solicitante=request.user)

    if request.method == 'POST':
        if pedido.status == 'PENDENTE':
            numero_pedido = pedido.id
            pedido.delete()
            messages.success(request, f"Pedido #{numero_pedido} foi excluído com sucesso.")
        else:
            messages.error(request, f"Este pedido não pode ser excluído pois seu status é '{pedido.get_status_display()}'.")

    return redirect('meus_pedidos')



# Apenas nutricionista pode ver a lista de pedidos pendentes
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def listar_pedidos_pendentes(request):

    pedidos = Pedido.objects.filter(
        status='PENDENTE', 
        itens__isnull=False
    ).select_related('solicitante__escola').distinct().order_by('data_pedido')
    
    context = {
        'pedidos': pedidos
    }
    return render(request, 'pedidos/listar_pedidos_pendentes.html', context)



# Apenas nutricionista pode ver os detalhes de um pedido
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def detalhe_pedido_nutricionista(request, pedido_id):
    pedido = get_object_or_404(Pedido.objects.select_related('solicitante', 'solicitante__escola'), id=pedido_id)
    context = {
        'pedido': pedido
    }
    return render(request, 'pedidos/detalhe_pedido_nutricionista.html', context)



# Apenas nutricionista pode aprovar um pedido
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def aprovar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if pedido.status != 'PENDENTE':
        messages.error(request, f"O pedido #{pedido.id} não está pendente.")
        return redirect('listar_pedidos_pendentes')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                
                produtos_ids = pedido.itens.values_list('produto_id', flat=True)
                produtos_com_estoque = TipoProduto.objects.filter(id__in=produtos_ids).annotate(
                    estoque_atual=Sum('lotes__quantidade_pacotes')
                )
                estoque_map = {p.id: p.estoque_atual or 0 for p in produtos_com_estoque}
                
                for item in pedido.itens.all():
                    quantidade_solicitada = item.quantidade
                    if quantidade_solicitada > estoque_map.get(item.produto.id, 0):
                        raise Exception(f"Estoque insuficiente para o produto '{item.produto.nome}'.")

                for item in pedido.itens.all():
                    quantidade_a_baixar = item.quantidade
                    lotes_disponiveis = Lote.objects.filter(
                        tipo_produto=item.produto, 
                        quantidade_pacotes__gt=0
                    ).order_by('data_validade')
                    
                    for lote in lotes_disponiveis:
                        if quantidade_a_baixar == 0:
                            break
                        
                        if lote.quantidade_pacotes >= quantidade_a_baixar:
                            lote.quantidade_pacotes -= quantidade_a_baixar
                            quantidade_a_baixar = 0
                        else:
                            quantidade_a_baixar -= lote.quantidade_pacotes
                            lote.quantidade_pacotes = 0
                            
                        lote.save()
                pedido.status = 'APROVADO'
                pedido.save()
                messages.success(request, f"Pedido #{pedido.id} aprovado e estoque atualizado!")

        except Exception as e:
            messages.error(request, f"Erro ao aprovar o pedido #{pedido.id}: {e}")

    return redirect('listar_pedidos_pendentes')



# Apenas nutricionista pode rejeitar um pedido
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def rejeitar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        motivo = request.POST.get('justificativa')
        
        pedido.status = 'REJEITADO'
        pedido.justificativa_rejeicao = motivo
        pedido.save()
        
        messages.success(request, f"Pedido #{pedido.id} rejeitado. Motivo registrado.")
        return redirect('listar_pedidos_pendentes')
    
    return redirect('detalhe_pedido_nutricionista', pedido_id=pedido.id)



# Apenas nutricionista pode verificar o historico de pedidos
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def historico_pedidos(request):
    pedidos_recentes = Pedido.objects.filter(
        Q(status='APROVADO') | Q(status='REJEITADO')
    ).select_related('solicitante__escola').order_by('-data_pedido')[:20]  

    context = {
        'pedidos': pedidos_recentes
    }
    
    return render(request, 'pedidos/historico_pedidos.html', context)



# Apenas nutricionista pode acessar a lista de pedidos pendentes
@login_required
@user_passes_test(lambda u: u.is_nutricionista())
def verificar_pedidos_pendentes(request):
    if not request.user.is_nutricionista():
        return JsonResponse({'count': 0, 'sucess': False})
    
    contagem = Pedido.objects.filter(status='PENDENTE').count()
    return JsonResponse({'count': contagem, 'success': True})



@login_required
@user_passes_test(lambda u: u.is_diretor())
def menu_diretor(request):
    return render(request, 'usuarios/menu_diretor.html')
   
   

# Apenas diretor pode excluir seu próprio pedido    
@login_required
@user_passes_test(lambda u: u.is_diretor())
def excluir_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, solicitante=request.user)

    if pedido.status == 'PENDENTE':
        numero = pedido.id
        pedido.delete() 
        messages.success(request, f"Pedido #{numero} excluído com sucesso.")
    else:
        messages.error(request, "Você só pode excluir pedidos que ainda estão pendentes.")

    return redirect('meus_pedidos')



# Apenas secretário pode acessar a lista de pedidos de carne
@login_required
@user_passes_test(lambda u: u.is_secretario())
def lista_compras_secretario(request):
    lista = Pedido.objects.exclude(observacao_carne__exact='') \
                          .exclude(observacao_carne__isnull=True) \
                          .filter(carne_comprada=False) \
                          .select_related('solicitante__escola')

    return render(request, 'pedidos/compras_diretas.html', {'lista': lista})




# Apenas secretário pode marcar o pedido de carne como comprado
@login_required
@user_passes_test(lambda u: u.is_secretario())
def marcar_carne_comprada(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    pedido.carne_comprada = True
    pedido.save()
    return redirect('lista_compras_secretario')



# Apenas secretário pode negar pedido de carne direta
@login_required
@user_passes_test(lambda u: u.is_secretario())
def negar_compra_direta(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        justificativa = request.POST.get('justificativa_secretario')
        
        pedido.justificativa_secretario = justificativa
        pedido.carne_comprada = True  # Marcando como "processado" para sair da lista de pendentes
        pedido.save()
        
        messages.warning(request, f"Solicitação do Pedido #{pedido.id} negada.")
        return redirect('lista_compras_secretario')

    return redirect('lista_compras_secretario')
